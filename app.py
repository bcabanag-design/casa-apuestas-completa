from flask import Flask, render_template, request, redirect, url_for, flash, make_response, g, session, jsonify
from casa_apuestas import CasaDeApuestas
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import os

# --- Helper para PyInstaller ---
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- Configuración Flask ---
template_folder = resource_path('templates')
app = Flask(__name__, template_folder=template_folder)
app.secret_key = "llave_secreta_para_local"

# --- Configuración DB ---
DB_FILENAME = 'casa_apuestas.db'
DB_PATH = resource_path(DB_FILENAME)

# --- Gestión DB ---
def get_casa():
    if 'casa' not in g:
        g.casa = CasaDeApuestas(DB_PATH)
    return g.casa

@app.teardown_appcontext
def teardown_casa(exception):
    casa = g.pop('casa', None)
    if casa is not None:
        casa.cerrar_conexion()

# --- Rutas de la Aplicación (Públicas) ---

@app.route('/')
def index():
    casa = get_casa()
    try:
        apostadores = casa.obtener_apostadores()
        partidas_abiertas = casa.obtener_partidas_abiertas()
        partidas_resueltas = casa.obtener_partidas_resueltas()
        apuestas_por_partida = {}
        for partida in partidas_abiertas:
            apuestas_por_partida[partida['id']] = casa.obtener_apuestas_partida(partida['id'])
    except Exception as e:
        print(f"Error index: {e}")
        flash("Error al cargar datos.", "error")
        apostadores, partidas_abiertas, partidas_resueltas = [], [], []
        apuestas_por_partida = {}

    return render_template('index.html', apostadores=apostadores, 
                           partidas_abiertas=partidas_abiertas, 
                           partidas_resueltas=partidas_resueltas, 
                           apuestas_por_partida=apuestas_por_partida)

@app.route('/add_apostador', methods=['POST'])
def add_apostador():
    casa = get_casa()
    nombre = request.form['nombre']
    try:
        saldo = float(request.form['saldo'])
        casa.registrar_apostador(nombre, saldo)
        flash(f"Apostador '{nombre}' registrado.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for('index'))

@app.route('/ajustar_saldo', methods=['POST'])
def ajustar_saldo():
    casa = get_casa()
    try:
        nombre = request.form['nombre_apostador']
        monto = float(request.form['monto'])
        casa.ajustar_saldo_apostador(nombre, monto)
        flash(f"Saldo ajustado para '{nombre}'.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for('index'))

@app.route('/crear_partida', methods=['POST'])
def crear_partida():
    casa = get_casa()
    try:
        casa.crear_partida(request.form['equipo1'], request.form['equipo2'])
        flash("Partida creada.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for('index', active_tab='partidas-abiertas'))

@app.route('/registrar_apuesta', methods=['POST'])
def registrar_apuesta():
    casa = get_casa()
    try:
        casa.registrar_apuesta(int(request.form['partida_id']), 
                               request.form['nombre_apostador'], 
                               float(request.form['monto']), 
                               int(request.form['equipo']))
        flash("Apuesta registrada.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for('index', active_tab='partidas-abiertas'))

@app.route('/resolver_partida', methods=['POST'])
def resolver_partida():
    casa = get_casa()
    try:
        casa.resolver_partida(int(request.form['partida_id']), int(request.form['equipo_ganador']))
        flash("Partida resuelta.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for('index', active_tab='partidas-resueltas'))

@app.route('/reportes')
def reportes():
    casa = get_casa()
    try:
        return render_template('reportes.html', 
                               apostadores=casa.obtener_balance_apostadores(), 
                               rentabilidad=casa.calcular_rentabilidad_total() or 0.0,
                               reporte_partidas=casa.obtener_reporte_partidas(),
                               reporte_apuestas_detallado=casa.obtener_reporte_apuestas_detallado())
    except Exception as e:
        flash(f"Error reportes: {e}", "error")
        return redirect(url_for('index'))

@app.route('/exportar_excel')
def exportar_excel():
    casa = get_casa()
    try:
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Hoja 1
        ws1 = wb.active
        ws1.title = "Balance Consolidado"
        ws1.append(['Apostador', 'Saldo Final', 'Total Apostado', 'Total Retornado', 'Ganancia Neta'])
        for cell in ws1[1]: cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align
        
        for a in casa.obtener_balance_apostadores():
            ws1.append([a.get('nombre'), a.get('saldo_final'), a.get('total_apostado'), a.get('total_ganado'), a.get('ganancia_neta')])
            
        # Hoja 2
        ws2 = wb.create_sheet("Ganancias Casa")
        ws2.append(['Partida', 'Ganador', 'Comisión (S/)'])
        for cell in ws2[1]: cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align
        
        for p in casa.obtener_reporte_partidas():
            ganador = p['nombre_equipo1'] if p['equipo_ganador'] == 1 else p['nombre_equipo2']
            ws2.append([f"{p['nombre_equipo1']} vs {p['nombre_equipo2']}", ganador, p['ganancia_casa']])
        ws2.append(['TOTAL', '', casa.calcular_rentabilidad_total() or 0.0])

        # Hoja 3
        ws3 = wb.create_sheet("Detalle Apuestas")
        ws3.append(['Partida ID', 'Apostador', 'Partida', 'Equipo Apostado', 'Monto', 'Cobrado', 'Neto', 'Resultado'])
        for cell in ws3[1]: cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align
        
        for ap in casa.obtener_reporte_apuestas_detallado():
            ws3.append([ap.get('partida_id'), ap.get('apostador'), ap.get('partida_nombre'), ap.get('equipo_apostado_nombre'),
                        ap.get('monto_apostado'), ap.get('monto_cobrado'), ap.get('ganancia_neta'), ap.get('resultado_texto')])

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = make_response(out.read())
        resp.headers["Content-Disposition"] = "attachment; filename=reporte_casa_apuestas.xlsx"
        resp.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return resp
    except Exception as e:
        print(f"Error Excel: {e}")
        flash(f"Error exportar: {e}", "error")
        return redirect(url_for('index'))

@app.route('/borrar_historial', methods=['POST'])
def borrar_historial():
    try:
        get_casa().borrar_partidas_resueltas()
        flash("Historial borrado.", "warning")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for('index', active_tab='partidas-resueltas'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)